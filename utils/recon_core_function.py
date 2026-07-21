"""
Generic two-dataset reconciliation logic.

Flow
----
1. User uploads a LEFT file and a RIGHT file (data expected on 'Sheet1', falls
   back to the first sheet if 'Sheet1' isn't found).
2. User optionally uploads a "Replace Values" correction file (data on
   'Sheet1', two columns: REPLACE / REPLACE WITH). Every occurrence of a
   REPLACE string found inside a mapping-column value is substituted with the
   REPLACE WITH string (case-insensitive; blank REPLACE WITH removes the
   text). Rules are applied in the order they appear in the file.
3. User picks one or more MAPPING column pairs (Left col <-> Right col).
   For each value: apply the replace rules first, THEN strip all special
   characters/spaces and uppercase, THEN concatenate the (ordered) mapping
   columns into a single matching key.
4. User picks one or more RECONCILIATION column pairs (Left col <-> Right
   col), e.g. IGST-vs-IGST, CGST-vs-CGST, SGST-vs-SGST. The label for each
   pair is auto-derived as "{left_col}-{right_col}".
   User also picks a reconciliation METHOD, applied to all pairs:
     - "sum"  : values are aggregated by the matching key on each side and
                the two totals are compared within tolerance.
     - "row"  : within each matching key, individual row values on the Left
                are matched one-to-one (in original row order, greedy
                first-fit) against unused Right row values within
                tolerance. Each reconciled row records which row (by
                original Excel row number) on the other side it matched.
   A difference within +/- tolerance counts as reconciled.
5. In the summary, the Unreconciled total for the RIGHT dataset is reported
   multiplied by -1 (sign convention: shortfall on the right shows negative).
"""
import re
import io
from dataclasses import dataclass
from typing import Dict, Any, List, Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, size=13)
NORMAL_FONT = Font(name="Arial", size=10)


# ---------------------------------------------------------------------------
# Cleaning / replace-values helpers
# ---------------------------------------------------------------------------

def clean_str(value):
    """Remove all non-alphanumeric characters, uppercase, strip spaces."""
    if pd.isna(value):
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(value)).upper()


def load_replace_rules_from_df(replace_df):
    """
    Build an ordered list of (pattern, replacement) tuples from a DataFrame
    with 'REPLACE' and 'REPLACE WITH' columns (case-insensitive column names).
    A blank/NaN REPLACE WITH means "remove". Returns [] if replace_df is None.
    """
    if replace_df is None:
        return []
    cols = {c.strip().upper(): c for c in replace_df.columns}
    replace_col = cols.get("REPLACE")
    replace_with_col = cols.get("REPLACE WITH")
    if replace_col is None or replace_with_col is None:
        raise ValueError("Replace Values data must have 'REPLACE' and 'REPLACE WITH' columns.")

    rules = []
    for _, row in replace_df.iterrows():
        pattern = row[replace_col]
        if pd.isna(pattern) or str(pattern).strip() == "":
            continue
        replacement = row[replace_with_col]
        replacement = "" if pd.isna(replacement) else str(replacement).strip()
        rules.append((str(pattern), replacement))
    return rules


def load_replace_rules(file_bytes):
    """
    Read the Replace Values correction file. Expects columns 'REPLACE' and
    'REPLACE WITH' on 'Sheet1' (or first sheet). Returns an ordered list of
    (pattern, replacement) tuples. A blank/NaN REPLACE WITH means "remove".
    """
    if file_bytes is None:
        return []
    sheet_name = get_target_sheet_name(file_bytes)
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    return load_replace_rules_from_df(df)


def apply_replace_rules(value, rules):
    """Apply ordered (pattern, replacement) rules to a single value, case-insensitive."""
    if pd.isna(value):
        return ""
    text = str(value)
    for pattern, replacement in rules:
        if pattern == "":
            continue
        text = re.sub(re.escape(pattern), replacement, text, flags=re.IGNORECASE)
    return text


def clean_mapping_value(value, rules):
    """Apply replace rules first, then strip special chars/spaces and uppercase."""
    replaced = apply_replace_rules(value, rules) if rules else value
    return clean_str(replaced)


# ---------------------------------------------------------------------------
# File / sheet reading helpers
# ---------------------------------------------------------------------------

def get_target_sheet_name(file_bytes):
    """Prefer a sheet literally named 'Sheet1' (case-insensitive); else first sheet."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    for name in xls.sheet_names:
        if name.strip().lower() == "sheet1":
            return name
    return xls.sheet_names[0]


def get_columns(file_bytes):
    """Return (sheet_name, list_of_column_names) for the target sheet."""
    sheet_name = get_target_sheet_name(file_bytes)
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, nrows=0)
    return sheet_name, list(df.columns)


def load_df(file_bytes, mapping_cols):
    """Read the target sheet, drop fully blank rows (based on mapping columns).

    Also stamps each row with its original Excel row number (assuming the
    header sits on row 1) in a 'Source Row' column, BEFORE dropping
    blank rows, so the numbers stay meaningful even after filtering.
    """
    sheet_name = get_target_sheet_name(file_bytes)
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    df.insert(0, "Source Row", df.index + 2)
    if mapping_cols:
        df = df[~df[mapping_cols].isna().all(axis=1)].copy()
    df.reset_index(drop=True, inplace=True)
    return df


def build_key(df, cols, rules=None):
    """Concatenate replace-corrected + cleaned values of the given columns into a key."""
    rules = rules or []

    def make_key(row):
        return "-".join(clean_mapping_value(row[c], rules) for c in cols)

    return df[cols].apply(make_key, axis=1)


def _numeric(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)


def is_numeric_column(df, col):
    """True if the column can be treated as numeric (after coercion, not all-NaN-when-source-had-values)."""
    coerced = pd.to_numeric(df[col], errors="coerce")
    non_null_source = df[col].notna().sum()
    non_null_coerced = coerced.notna().sum()
    if non_null_source == 0:
        return True
    return non_null_coerced >= non_null_source * 0.95  # allow a few stray non-numeric cells


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

def reconcile_pair(left_df, right_df, left_col, right_col, tolerance):
    l_vals = _numeric(left_df[left_col])
    r_vals = _numeric(right_df[right_col])
    agg_left = l_vals.groupby(left_df["Key"]).sum()
    agg_right = r_vals.groupby(right_df["Key"]).sum()
    all_keys = set(agg_left.index) | set(agg_right.index)

    status_map = {}
    for k in all_keys:
        v_l = agg_left.get(k, 0)
        v_r = agg_right.get(k, 0)
        diff = v_l - v_r
        status_map[k] = "Reconciled" if abs(diff) <= tolerance else "Unreconciled"

    return status_map, agg_left, agg_right


def _build_pair_tables_sum(left_df, right_df, left_col, right_col, pair_label, tolerance):
    """Aggregate-by-key ('Sum Value') reconciliation method."""
    status_map, agg_left, agg_right = reconcile_pair(left_df, right_df, left_col, right_col, tolerance)

    left_out = left_df.copy()
    right_out = right_df.copy()

    left_out["MAPPED (Matching) Key"] = left_out["Key"]
    right_out["MAPPED (Matching) Key"] = right_out["Key"]

    left_out[f"Aggregated {pair_label} - Left"] = left_out["Key"].map(agg_left).fillna(0)
    left_out[f"Aggregated {pair_label} - Right"] = left_out["Key"].map(agg_right).fillna(0)
    right_out[f"Aggregated {pair_label} - Left"] = right_out["Key"].map(agg_left).fillna(0)
    right_out[f"Aggregated {pair_label} - Right"] = right_out["Key"].map(agg_right).fillna(0)

    left_out[f"{pair_label} Difference"] = left_out[f"Aggregated {pair_label} - Left"] - left_out[f"Aggregated {pair_label} - Right"]
    right_out[f"{pair_label} Difference"] = right_out[f"Aggregated {pair_label} - Left"] - right_out[f"Aggregated {pair_label} - Right"]

    status_label = {"Reconciled": "MAPPED (Matching)", "Unreconciled": "UNRECONCILED"}
    left_out[f"{pair_label} Reconciliation Status"] = left_out["Key"].map(status_map).map(status_label)
    right_out[f"{pair_label} Reconciliation Status"] = right_out["Key"].map(status_map).map(status_label)

    total_left = _numeric(left_df[left_col]).sum()
    total_right = _numeric(right_df[right_col]).sum()
    recon_left = _numeric(left_out.loc[left_out[f"{pair_label} Reconciliation Status"] == "MAPPED (Matching)", left_col]).sum()
    recon_right = _numeric(right_out.loc[right_out[f"{pair_label} Reconciliation Status"] == "MAPPED (Matching)", right_col]).sum()
    unrecon_left = _numeric(left_out.loc[left_out[f"{pair_label} Reconciliation Status"] == "UNRECONCILED", left_col]).sum()
    unrecon_right = _numeric(right_out.loc[right_out[f"{pair_label} Reconciliation Status"] == "UNRECONCILED", right_col]).sum()

    # Sign convention: Unreconciled total of the RIGHT dataset is reported negative.
    unrecon_right_signed = unrecon_right * -1

    summary = {
        f"Total {pair_label} as per Left ({left_col})": total_left,
        f"Total {pair_label} as per Right ({right_col})": total_right,
        "Difference (Left - Right)": total_left - total_right,
        f"Reconciled {pair_label} - Left": recon_left,
        f"Reconciled {pair_label} - Right": recon_right,
        f"Unreconciled {pair_label} - Left": unrecon_left,
        f"Unreconciled {pair_label} - Right": unrecon_right_signed,
        "Matching tolerance": tolerance,
        "Reconciliation Method": "Sum Value",
    }

    reconciled_left = left_out[left_out[f"{pair_label} Reconciliation Status"] == "MAPPED (Matching)"].copy()
    reconciled_right = right_out[right_out[f"{pair_label} Reconciliation Status"] == "MAPPED (Matching)"].copy()
    unreconciled_left = left_out[left_out[f"{pair_label} Reconciliation Status"] == "UNRECONCILED"].copy()
    unreconciled_right = right_out[right_out[f"{pair_label} Reconciliation Status"] == "UNRECONCILED"].copy()

    return {
        "summary": summary,
        "reconciled_left": reconciled_left,
        "reconciled_right": reconciled_right,
        "unreconciled_left": unreconciled_left,
        "unreconciled_right": unreconciled_right,
        "status_map": status_map,
        "totals": {
            "left_col": left_col,
            "right_col": right_col,
            "method": "Sum Value",
            "tolerance": tolerance,
            "total_left": total_left,
            "total_right": total_right,
            "difference": total_left - total_right,
            "reconciled_left": recon_left,
            "reconciled_right": recon_right,
            "unreconciled_left": unrecon_left,
            "unreconciled_right_signed": unrecon_right_signed,
        },
    }


def _match_rows_within_key(left_values, right_values, tolerance):
    """
    Greedy first-fit, top-to-bottom matching of two lists of numeric values
    (both already restricted to a single matching key, in original row
    order). Returns a list, same length as left_values, of either the
    matched index into right_values or None.
    """
    used_right = [False] * len(right_values)
    matches = []
    for lv in left_values:
        matched_ri = None
        for ri, rv in enumerate(right_values):
            if used_right[ri]:
                continue
            if abs(lv - rv) <= tolerance:
                matched_ri = ri
                used_right[ri] = True
                break
        matches.append(matched_ri)
    return matches


def _build_pair_tables_row(left_df, right_df, left_col, right_col, pair_label, tolerance):
    """
    Row-level ('Row Value') reconciliation method: within each matching key,
    each Left row is matched 1:1 (in original order, greedy first-fit)
    against an unused Right row whose value is within tolerance. Reconciled
    rows record which row (by original Excel row number) they matched on the
    other side.
    """
    left_out = left_df.copy().reset_index(drop=True)
    right_out = right_df.copy().reset_index(drop=True)

    status_col = f"{pair_label} Reconciliation Status"
    matched_left_ref_col = f"{pair_label} Matched Row (Right)"
    matched_right_ref_col = f"{pair_label} Matched Row (Left)"

    left_out["MAPPED (Matching) Key"] = left_out["Key"]
    right_out["MAPPED (Matching) Key"] = right_out["Key"]
    left_out[status_col] = "UNRECONCILED"
    right_out[status_col] = "UNRECONCILED"
    left_out[matched_left_ref_col] = pd.NA
    right_out[matched_right_ref_col] = pd.NA

    l_values_all = _numeric(left_out[left_col])
    r_values_all = _numeric(right_out[right_col])

    left_groups = left_out.groupby("Key").indices
    right_groups = right_out.groupby("Key").indices
    all_keys = set(left_groups.keys()) | set(right_groups.keys())

    for key in all_keys:
        l_positions = sorted(left_groups.get(key, []))
        r_positions = sorted(right_groups.get(key, []))
        l_vals = [l_values_all.iloc[p] for p in l_positions]
        r_vals = [r_values_all.iloc[p] for p in r_positions]

        matches = _match_rows_within_key(l_vals, r_vals, tolerance)

        for local_li, local_ri in enumerate(matches):
            global_l_pos = l_positions[local_li]
            if local_ri is not None:
                global_r_pos = r_positions[local_ri]
                left_out.at[global_l_pos, status_col] = "MAPPED (Matching)"
                right_out.at[global_r_pos, status_col] = "MAPPED (Matching)"
                left_out.at[global_l_pos, matched_left_ref_col] = right_out.at[global_r_pos, "Source Row"]
                right_out.at[global_r_pos, matched_right_ref_col] = left_out.at[global_l_pos, "Source Row"]

    total_left = l_values_all.sum()
    total_right = r_values_all.sum()
    recon_left = l_values_all[left_out[status_col] == "MAPPED (Matching)"].sum()
    recon_right = r_values_all[right_out[status_col] == "MAPPED (Matching)"].sum()
    unrecon_left = l_values_all[left_out[status_col] == "UNRECONCILED"].sum()
    unrecon_right = r_values_all[right_out[status_col] == "UNRECONCILED"].sum()

    # Sign convention: Unreconciled total of the RIGHT dataset is reported negative.
    unrecon_right_signed = unrecon_right * -1

    summary = {
        f"Total {pair_label} as per Left ({left_col})": total_left,
        f"Total {pair_label} as per Right ({right_col})": total_right,
        "Difference (Left - Right)": total_left - total_right,
        f"Reconciled {pair_label} - Left": recon_left,
        f"Reconciled {pair_label} - Right": recon_right,
        f"Unreconciled {pair_label} - Left": unrecon_left,
        f"Unreconciled {pair_label} - Right": unrecon_right_signed,
        "Matching tolerance": tolerance,
        "Reconciliation Method": "Row Value",
    }

    reconciled_left = left_out[left_out[status_col] == "MAPPED (Matching)"].copy()
    reconciled_right = right_out[right_out[status_col] == "MAPPED (Matching)"].copy()
    unreconciled_left = left_out[left_out[status_col] == "UNRECONCILED"].copy()
    unreconciled_right = right_out[right_out[status_col] == "UNRECONCILED"].copy()

    return {
        "summary": summary,
        "reconciled_left": reconciled_left,
        "reconciled_right": reconciled_right,
        "unreconciled_left": unreconciled_left,
        "unreconciled_right": unreconciled_right,
        "totals": {
            "left_col": left_col,
            "right_col": right_col,
            "method": "Row Value",
            "tolerance": tolerance,
            "total_left": total_left,
            "total_right": total_right,
            "difference": total_left - total_right,
            "reconciled_left": recon_left,
            "reconciled_right": recon_right,
            "unreconciled_left": unrecon_left,
            "unreconciled_right_signed": unrecon_right_signed,
        },
    }


def build_pair_tables(left_df, right_df, left_col, right_col, pair_label, tolerance, method="sum"):
    """Dispatch to the Sum-Value or Row-Value reconciliation method."""
    if method == "row":
        return _build_pair_tables_row(left_df, right_df, left_col, right_col, pair_label, tolerance)
    return _build_pair_tables_sum(left_df, right_df, left_col, right_col, pair_label, tolerance)


def run_reconciliation(left_bytes, right_bytes, mapping_pairs, recon_pairs, tolerance, replace_rules=None, method="sum"):
    """
    mapping_pairs: list of (left_col, right_col) tuples
    recon_pairs: list of (label, left_col, right_col) tuples
    replace_rules: list of (pattern, replacement) tuples, applied to mapping columns
    method: "sum" (aggregate-by-key totals) or "row" (row-by-row 1:1 matching)
    """
    replace_rules = replace_rules or []
    left_map_cols = [p[0] for p in mapping_pairs]
    right_map_cols = [p[1] for p in mapping_pairs]

    left_df = load_df(left_bytes, left_map_cols)
    right_df = load_df(right_bytes, right_map_cols)

    left_df["Key"] = build_key(left_df, left_map_cols, replace_rules)
    right_df["Key"] = build_key(right_df, right_map_cols, replace_rules)

    results = {}
    for label, left_col, right_col in recon_pairs:
        results[label] = build_pair_tables(left_df, right_df, left_col, right_col, label, tolerance, method=method)

    return {
        "left_df": left_df,
        "right_df": right_df,
        "pair_results": results,
        "recon_pairs": recon_pairs,
        "method": method,
    }


# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------

INVALID_SHEET_CHARS = r'[\[\]\:\*\?/\\]'


def safe_sheet_name(name, suffix="", used_names=None):
    base = re.sub(INVALID_SHEET_CHARS, "", str(name)).strip()
    if not base:
        base = "Sheet"
    max_len = 31 - len(suffix)
    candidate = (base[:max_len] + suffix) if max_len > 0 else suffix[:31]
    if used_names is not None:
        i = 1
        while candidate in used_names:
            trim = 31 - len(suffix) - len(f"_{i}")
            candidate = f"{base[:max(trim,1)]}_{i}{suffix}"
            i += 1
        used_names.add(candidate)
    return candidate


def _autofit_columns(ws, max_width=45):
    for col_cells in ws.columns:
        length = 0
        col_letter = None
        for cell in col_cells:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
            except Exception:
                val_len = 0
            length = max(length, val_len)
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def _write_df_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    if df.empty:
        ws.append(["No records"])
        ws["A1"].font = NORMAL_FONT
        return ws

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    numeric_cols = set(df.select_dtypes(include="number").columns)
    for _, row in df.iterrows():
        clean_row = [None if pd.isna(v) else v for v in row.values]
        ws.append(clean_row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row):
            cell.font = NORMAL_FONT
            col_name = df.columns[idx]
            if col_name in numeric_cols:
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    _autofit_columns(ws)
    return ws


def _write_summary_sheet(wb, sheet_name, label, summary_dict):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"{label} Reconciliation Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A3"] = "Description"
    ws["B3"] = "Amount"
    for col in ("A3", "B3"):
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL

    row = 4
    for desc, amount in summary_dict.items():
        ws.cell(row=row, column=1, value=desc).font = NORMAL_FONT
        cell = ws.cell(row=row, column=2, value=amount)
        cell.font = NORMAL_FONT
        if desc not in ("Matching tolerance", "Reconciliation Method"):
            cell.number_format = "#,##0.00"
        row += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    return ws


def build_report_workbook(recon_output):
    """Builds the full multi-sheet reconciliation report workbook (in-memory)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_names = set()
    for label, res in recon_output["pair_results"].items():
        _write_summary_sheet(wb, safe_sheet_name(label, " Summary", used_names), label, res["summary"])
        _write_df_sheet(wb, safe_sheet_name(label, " Rec-L", used_names), res["reconciled_left"])
        _write_df_sheet(wb, safe_sheet_name(label, " Rec-R", used_names), res["reconciled_right"])
        _write_df_sheet(wb, safe_sheet_name(label, " Unrec-L", used_names), res["unreconciled_left"])
        _write_df_sheet(wb, safe_sheet_name(label, " Unrec-R", used_names), res["unreconciled_right"])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ---------------------------------------------------------------------------
# Reusable, UI-independent API — DataFrames in, DataFrames + Excel bytes out
# ---------------------------------------------------------------------------

@dataclass
class ReconciliationResult:
    """
    excel_bytes : the full multi-sheet reconciliation report workbook, ready
                   to write to disk or hand to a download button.
    summary     : one tidy row per reconciliation column pair with the key
                   totals (Total/Reconciled/Unreconciled Left & Right, etc).
    details     : {pair_label: {"summary": {...}, "reconciled_left": df,
                   "reconciled_right": df, "unreconciled_left": df,
                   "unreconciled_right": df, "totals": {...}}} — full
                   row-level detail for every pair.
    left_df / right_df : the working copies of the input data actually used
                   for matching (Source Row + Key columns added).
    """
    excel_bytes: bytes
    summary: pd.DataFrame
    details: Dict[str, Dict[str, Any]]
    left_df: pd.DataFrame
    right_df: pd.DataFrame


def reconcile_dataframes(
    left_df: pd.DataFrame,
    right_df: pd.DataFrame,
    mapping_cols_left: List[str],
    mapping_cols_right: List[str],
    recon_cols_left: List[str],
    recon_cols_right: List[str],
    tolerance: float = 1,
    method: str = "sum",
    replace_df: Optional[pd.DataFrame] = None,
) -> ReconciliationResult:
    """
    Reusable, non-UI reconciliation entry point. Give it two already-loaded
    DataFrames and column lists; get back an Excel report plus pandas
    DataFrames for the summary and full detail.

    Parameters
    ----------
    left_df, right_df : the two datasets to reconcile.
    mapping_cols_left, mapping_cols_right : equal-length lists of column
        names (paired by position) used to build the matching key. Values
        are replace-corrected (if replace_df given), then stripped of all
        special characters/spaces and uppercased, then concatenated.
    recon_cols_left, recon_cols_right : equal-length lists of NUMERIC column
        names (paired by position) to reconcile, e.g.
        recon_cols_left=['IGST','CGST','SGST'], recon_cols_right=['IGST','CGST','SGST'].
        Each pair is labeled "{left_col}-{right_col}" automatically.
    tolerance : a difference within +/- tolerance counts as reconciled.
        Applies to every reconciliation pair.
    method : "sum" — aggregate each side by matching key and compare totals.
             "row" — within each matching key, match Left rows 1:1 (top to
                     bottom, greedy first-fit) against unused Right rows
                     whose value is within tolerance; matched rows record
                     which row on the other side they paired with.
    replace_df : optional DataFrame with 'REPLACE' / 'REPLACE WITH' columns,
        applied (in row order, case-insensitive) to mapping-column values
        before cleaning. Pass None to skip.

    Returns
    -------
    ReconciliationResult(excel_bytes, summary, details, left_df, right_df)

    Notes
    -----
    In both the `summary` DataFrame and every per-pair summary, the
    Unreconciled total for the RIGHT dataset is reported multiplied by -1
    (shortfall on the right shows negative).
    """
    if len(mapping_cols_left) != len(mapping_cols_right):
        raise ValueError("mapping_cols_left and mapping_cols_right must be the same length.")
    if len(recon_cols_left) != len(recon_cols_right):
        raise ValueError("recon_cols_left and recon_cols_right must be the same length.")
    if len(recon_cols_left) == 0:
        raise ValueError("At least one reconciliation column pair is required.")
    if method not in ("sum", "row"):
        raise ValueError("method must be 'sum' or 'row'.")

    replace_rules = load_replace_rules_from_df(replace_df)

    # Working copies: stamp a Source Row reference (the caller's original
    # index) before any filtering/reindexing, so matches can be traced back.
    left_work = left_df.copy()
    right_work = right_df.copy()
    left_work.insert(0, "Source Row", left_work.index)
    right_work.insert(0, "Source Row", right_work.index)

    if mapping_cols_left:
        left_work = left_work[~left_work[mapping_cols_left].isna().all(axis=1)].copy()
    if mapping_cols_right:
        right_work = right_work[~right_work[mapping_cols_right].isna().all(axis=1)].copy()
    left_work.reset_index(drop=True, inplace=True)
    right_work.reset_index(drop=True, inplace=True)

    left_work["Key"] = build_key(left_work, mapping_cols_left, replace_rules)
    right_work["Key"] = build_key(right_work, mapping_cols_right, replace_rules)

    recon_pairs = [
        (f"{lcol}-{rcol}", lcol, rcol)
        for lcol, rcol in zip(recon_cols_left, recon_cols_right)
    ]

    details = {}
    for label, lcol, rcol in recon_pairs:
        details[label] = build_pair_tables(left_work, right_work, lcol, rcol, label, tolerance, method=method)

    recon_output = {
        "left_df": left_work,
        "right_df": right_work,
        "pair_results": details,
        "recon_pairs": recon_pairs,
        "method": method,
    }
    excel_bio = build_report_workbook(recon_output)

    summary_rows = []
    for label, res in details.items():
        t = res["totals"]
        summary_rows.append({
            "Pair": label,
            "Left Column": t["left_col"],
            "Right Column": t["right_col"],
            "Method": t["method"],
            "Tolerance": t["tolerance"],
            "Total Left": t["total_left"],
            "Total Right": t["total_right"],
            "Difference (Left-Right)": t["difference"],
            "Reconciled Left": t["reconciled_left"],
            "Reconciled Right": t["reconciled_right"],
            "Unreconciled Left": t["unreconciled_left"],
            "Unreconciled Right (signed)": t["unreconciled_right_signed"],
        })
    summary_df = pd.DataFrame(summary_rows)

    return ReconciliationResult(
        excel_bytes=excel_bio.getvalue(),
        summary=summary_df,
        details=details,
        left_df=left_work,
        right_df=right_work,
    )
