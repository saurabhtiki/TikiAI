"""
Core reconciliation logic for GST 3B vs Purchase Register reconciliation.
Handles IGST, CGST, SGST reconciliation separately, builds report workbook,
and produces an ITC-Availability-updated copy of the GST3B workbook.
"""
import re
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TAXES = ["IGST", "CGST", "SGST"]

GST3B_SHEET = "B2B AND B2BA"
GST3B_HEADER_ROW_0INDEX = 3      # pandas header= (0-indexed) -> row 4 in Excel
GST3B_DATA_START_EXCEL_ROW = 5   # first data row in the raw worksheet

PR_SHEET = "WORKING"
PR_HEADER_ROW_0INDEX = 1         # pandas header= (0-indexed) -> row 2 in Excel

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, size=13)
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)


def clean_str(value):
    """Remove all non-alphanumeric characters, uppercase, strip spaces."""
    if pd.isna(value):
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(value)).upper()


def load_gst3b_df(file_bytes):
    """Read the B2B AND B2BA sheet into a DataFrame, keeping only real data rows."""
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=GST3B_SHEET, header=GST3B_HEADER_ROW_0INDEX)
    # Keep only rows that actually have a GSTIN (drops trailing blank rows)
    df = df[df["GSTIN of supplier"].notna()].copy()
    df["Clean GSTIN"] = df["GSTIN of supplier"].apply(clean_str)
    df["Clean Invoice Number"] = df["Invoice number"].apply(clean_str)
    df["Key"] = df["Clean GSTIN"] + "-" + df["Clean Invoice Number"]
    for tax in TAXES:
        df[tax] = pd.to_numeric(df[tax], errors="coerce").fillna(0)
    return df


def load_purchase_register_df(file_bytes):
    """Read the WORKING sheet into a DataFrame, keeping only real data rows."""
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=PR_SHEET, header=PR_HEADER_ROW_0INDEX)
    # Keep only rows that actually have an Invoice No. (drops trailing blank/total rows)
    df = df[df["Invoice No."].notna()].copy()
    df["Clean GSTIN"] = df["GST Reg. No."].apply(clean_str)
    df["Clean Invoice Number"] = df["Comm.  Inv. No."].apply(clean_str)
    df["Key"] = df["Clean GSTIN"] + "-" + df["Clean Invoice Number"]
    for tax in TAXES:
        df[tax] = pd.to_numeric(df[tax], errors="coerce").fillna(0)
    return df


def reconcile_tax(pr_df, gst3b_df, tax, tolerance):
    """
    Reconcile a single tax type (IGST/CGST/SGST) between Purchase Register and GST3B
    on an aggregated-by-key basis. Returns:
      - status_map: dict key -> 'Reconciled' / 'Unreconciled'
      - agg_pr: Series key -> aggregated PR tax amount
      - agg_3b: Series key -> aggregated 3B tax amount
    """
    agg_pr = pr_df.groupby("Key")[tax].sum()
    agg_3b = gst3b_df.groupby("Key")[tax].sum()
    all_keys = set(agg_pr.index) | set(agg_3b.index)

    status_map = {}
    for k in all_keys:
        v_pr = agg_pr.get(k, 0)
        v_3b = agg_3b.get(k, 0)
        diff = v_pr - v_3b
        status_map[k] = "Reconciled" if abs(diff) <= tolerance else "Unreconciled"

    return status_map, agg_pr, agg_3b


def build_tax_tables(pr_df, gst3b_df, tax, tolerance):
    """
    Build the per-tax augmented PR and GST3B dataframes (with mapping/status columns added),
    plus a summary dict, mirroring the structure of the reference reconciliation output file.
    """
    status_map, agg_pr, agg_3b = reconcile_tax(pr_df, gst3b_df, tax, tolerance)

    pr_out = pr_df.copy()
    gst3b_out = gst3b_df.copy()

    pr_out["MAPPED (Matching)- GST-Invoice"] = pr_out["Key"]
    gst3b_out["MAPPED (Matching)- GST-Invoice"] = gst3b_out["Key"]

    pr_out[f"Aggregated {tax} - Purchase Register"] = pr_out["Key"].map(agg_pr).fillna(0)
    pr_out[f"Aggregated {tax} - 3B GST"] = pr_out["Key"].map(agg_3b).fillna(0)
    gst3b_out[f"Aggregated {tax} - Purchase Register"] = gst3b_out["Key"].map(agg_pr).fillna(0)
    gst3b_out[f"Aggregated {tax} - 3B GST"] = gst3b_out["Key"].map(agg_3b).fillna(0)
    
    pr_out[f"{tax} Difference"] = pr_out[f"Aggregated {tax} - Purchase Register"] - pr_out[f"Aggregated {tax} - 3B GST"]
    gst3b_out[f"{tax} Difference"] = gst3b_out[f"Aggregated {tax} - Purchase Register"] - gst3b_out[f"Aggregated {tax} - 3B GST"]

    status_label = {"Reconciled": "MAPPED (Matching)", "Unreconciled": "UNRECONCILED"}
    pr_out[f"{tax} Reconciliation Status"] = pr_out["Key"].map(status_map).map(status_label)
    gst3b_out[f"{tax} Reconciliation Status"] = gst3b_out["Key"].map(status_map).map(status_label)

    total_pr = pr_df[tax].sum()
    total_3b = gst3b_df[tax].sum()
    recon_pr = pr_out.loc[pr_out[f"{tax} Reconciliation Status"] == "MAPPED (Matching)", tax].sum()
    recon_3b = gst3b_out.loc[gst3b_out[f"{tax} Reconciliation Status"] == "MAPPED (Matching)", tax].sum()
    unrecon_pr = pr_out.loc[pr_out[f"{tax} Reconciliation Status"] == "UNRECONCILED", tax].sum()
    unrecon_3b = gst3b_out.loc[gst3b_out[f"{tax} Reconciliation Status"] == "UNRECONCILED", tax].sum()*-1
    #dunreconciled difference of left dataset should be negative
    summary = {
        "Total {0} as per Purchase Register".format(tax): total_pr,
        "Total {0} as per 3B GST".format(tax): total_3b,
        "Difference (Purchase Register - 3B GST)": total_pr - total_3b,
        "Reconciled {0} - Purchase Register".format(tax): recon_pr,
        "Reconciled {0} - 3B GST".format(tax): recon_3b,
        "Unreconciled {0} - Purchase Register".format(tax): unrecon_pr,
        "Unreconciled {0} - 3B GST".format(tax): unrecon_3b,
        "Matching tolerance": tolerance,
    }

    reconciled_books = pr_out[pr_out[f"{tax} Reconciliation Status"] == "MAPPED (Matching)"].copy()
    reconciled_3b = gst3b_out[gst3b_out[f"{tax} Reconciliation Status"] == "MAPPED (Matching)"].copy()
    unreconciled_books = pr_out[pr_out[f"{tax} Reconciliation Status"] == "UNRECONCILED"].copy()
    unreconciled_3b = gst3b_out[gst3b_out[f"{tax} Reconciliation Status"] == "UNRECONCILED"].copy()

    return {
        "summary": summary,
        "reconciled_books": reconciled_books,
        "reconciled_3b": reconciled_3b,
        "unreconciled_books": unreconciled_books,
        "unreconciled_3b": unreconciled_3b,
        "status_map": status_map,
    }


def run_full_reconciliation(gst3b_bytes, pr_bytes, tolerance):
    """
    Runs reconciliation for IGST, CGST, SGST. Returns a dict keyed by tax
    with the tables described above, plus the raw dataframes for later use
    (e.g. ITC Availability update).
    """
    gst3b_df = load_gst3b_df(gst3b_bytes)
    pr_df = load_purchase_register_df(pr_bytes)

    results = {}
    for tax in TAXES:
        results[tax] = build_tax_tables(pr_df, gst3b_df, tax, tolerance)

    return {
        "gst3b_df": gst3b_df,
        "pr_df": pr_df,
        "tax_results": results,
    }


# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------

def _autofit_columns(ws, max_width=45):
    for col_cells in ws.columns:
        length = 0
        col_letter = None
        for cell in col_cells:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
            except Exception:
                val_len = 0
            length = max(length, val_len)
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def _write_df_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    if df.empty:
        ws.append(["No records"])
        ws["A1"].font = NORMAL_FONT
        return ws

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    numeric_cols = set(df.select_dtypes(include="number").columns)
    for _, row in df.iterrows():
        ws.append(list(row.values))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row):
            cell.font = NORMAL_FONT
            col_name = df.columns[idx]
            if col_name in numeric_cols:
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    _autofit_columns(ws)
    return ws


def _write_summary_sheet(wb, sheet_name, tax, summary_dict):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"{tax} Reconciliation Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A3"] = "Description"
    ws["B3"] = "Amount"
    for col in ("A3", "B3"):
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL

    row = 4
    for desc, amount in summary_dict.items():
        ws.cell(row=row, column=1, value=desc).font = NORMAL_FONT
        cell = ws.cell(row=row, column=2, value=amount)
        cell.font = NORMAL_FONT
        if desc != "Matching tolerance":
            cell.number_format = "#,##0.00"
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    return ws


def build_report_workbook(recon_output):
    """Builds the full multi-sheet reconciliation report workbook (in-memory)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tax_results = recon_output["tax_results"]
    for tax in TAXES:
        res = tax_results[tax]
        _write_summary_sheet(wb, f"{tax} Summary", tax, res["summary"])
        _write_df_sheet(wb, f"{tax} Reconciled Books", res["reconciled_books"])
        _write_df_sheet(wb, f"{tax} Reconciled 3B", res["reconciled_3b"])
        _write_df_sheet(wb, f"{tax} Unreconciled Books", res["unreconciled_books"])
        _write_df_sheet(wb, f"{tax} Unreconciled 3B", res["unreconciled_3b"])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_updated_gst3b_workbook(gst3b_bytes, recon_output):
    """
    Loads the ORIGINAL GST3B workbook (preserving every sheet, formatting, and
    data validation) and updates only the 'ITC Availability' column on the
    'B2B AND B2BA' sheet according to the ACCEPT rule:
      - IGST == 0 and CGST reconciled for that key -> ACCEPT
      - CGST == 0 and IGST reconciled for that key -> ACCEPT
      - existing ACCEPT values are preserved
    """
    igst_status = recon_output["tax_results"]["IGST"]["status_map"]
    cgst_status = recon_output["tax_results"]["CGST"]["status_map"]

    wb = openpyxl.load_workbook(io.BytesIO(gst3b_bytes))
    ws = wb[GST3B_SHEET]

    # Column positions (1-indexed) based on header row layout:
    # A: ITC Availability, B: GSTIN of supplier, D: Invoice number,
    # K: IGST, L: CGST
    COL_ITC = 1
    COL_GSTIN = 2
    COL_INVOICE = 4
    COL_IGST = 11
    COL_CGST = 12

    max_row = ws.max_row
    for r in range(GST3B_DATA_START_EXCEL_ROW, max_row + 1):
        gstin = ws.cell(row=r, column=COL_GSTIN).value
        if gstin is None or str(gstin).strip() == "":
            continue
        invoice = ws.cell(row=r, column=COL_INVOICE).value
        igst_val = ws.cell(row=r, column=COL_IGST).value or 0
        cgst_val = ws.cell(row=r, column=COL_CGST).value or 0

        key = clean_str(gstin) + "-" + clean_str(invoice)

        current = ws.cell(row=r, column=COL_ITC).value
        new_accept = False
        try:
            igst_val_f = float(igst_val)
        except (TypeError, ValueError):
            igst_val_f = 0.0
        try:
            cgst_val_f = float(cgst_val)
        except (TypeError, ValueError):
            cgst_val_f = 0.0

        if igst_val_f == 0 and cgst_status.get(key) == "Reconciled":
            new_accept = True
        if cgst_val_f == 0 and igst_status.get(key) == "Reconciled":
            new_accept = True

        if new_accept or (isinstance(current, str) and current.strip().upper() == "ACCEPT"):
            ws.cell(row=r, column=COL_ITC).value = "ACCEPT"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
